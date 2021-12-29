from django.core.management.base import BaseCommand

from openpyxl import load_workbook
from decimal import Decimal
from products.models import (
    Material,
    BathroomProducts,
    DecorationsProducts,
    FabricTextileProducts,
    FurnitureProducts,
    HardwareToolProducts,
    HomeApplianceProducts,
    KitchenProducts,
    LandscapeProducts,
    LightProducts,
    RugsMatFloorProducts,
    SecurityProtectionProducts,
    ProductImage,
    Brand,
)


class Command(BaseCommand):
    help = "Input data from excel to db"

    def add_arguments(self, parser):
        parser.add_argument("excel", type=str, help="string path of the excel file")

    def handle(self, *args, **kwargs):
        excel = kwargs["excel"]

        workbook = load_workbook(filename=excel)
        sheet = workbook.active

        products = []

        for row in sheet.iter_rows(
            min_row=4, max_row=143, min_col=2, max_col=16, values_only=True
        ):
            product_brand = Brand.objects.get(name=row[5])

            if row[3] == "Finish Materials":
                product_category = Material

            product = {
                "category": product_category,
                "name": str(row[4]) + "-" + str(row[1]) + str(row[2]),
                "brand": product_brand,
                "description": row[6] + row[7],
                "price": Decimal(row[14]) if row[14] else Decimal("500.00"),
                "available": True if row[12] == "In stock" else False,
                "material_place": "INDOOR",
                "material_category": "WALL",
            }
            products.append(product)

        for product in products:
            product_cat = product["category"]
            product.pop("category")
            product_cat.objects.create(**product)
